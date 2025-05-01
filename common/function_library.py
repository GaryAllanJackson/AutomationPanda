import datetime
import time
from ftplib import print_line

# from ftplib import print_line

import openpyxl
import requests
from selenium.webdriver.common.by import By

from common import variables


class Functions:
    def __init__(self, driver):
        self.driver = driver
        self.log_file_name = variables.log_file_name
        self.log_sheet_name = variables.log_sheet_name


    def navigate(self, page_url, description):
        self.driver.get(page_url)
        time.sleep(3)
        self.log_equal_action("navigate", page_url, self.driver.current_url, description)
        assert page_url == self.driver.current_url, "Navigation Failed!"

    def perform_click(self, accessor_type, accessor, expected, actual, description):
        element = self.get_element(accessor_type, accessor)
        element.click()
        # print(f"expected = {expected}")
        if expected != "n/a":
            time.sleep(3)
            if actual.lower() == variables.get_current_url_in_action_method:
                actual = self.driver.current_url
        self.log_equal_action("click", expected, actual, description)
        assert expected == actual, "Click Failed!"

    def perform_send_key(self, accessor_type, accessor, send_text, expected, actual, description):
        element = self.get_element(accessor_type, accessor)
        element.send_keys(send_text)
        #in case the actual value needs to be retrieved from the element where text was sent
        if actual == "get_text":
            actual = self.get_element_text_alt(element)

        self.log_equal_action("Send Keys", expected, actual, description)
        assert expected == actual, "Click Failed!"

    # This method returns the element text to the calling method
    # based on the type of element passed in
    def get_element_text(self, accessor_type, accessor):
        element = self.get_element(accessor_type, accessor)
        print("Retrieving text for element: ", element.get_attribute("outerHTML"))
        el_html = element.get_attribute("outerHTML")
        returnValue = ""
        if el_html.find("<input") > -1:
            returnValue = element.get_attribute("value")
        else:
            try:
                returnValue = element.text
            except:
                returnValue = element.get_attribute("innerText")

        return returnValue

    def get_element_text_alt(self, element):

        # print("Retrieving text for element: ", element.get_attribute("outerHTML"))
        el_html = element.get_attribute("outerHTML")
        returnValue = ""
        if el_html.find("<input") > -1:
            returnValue = element.get_attribute("value")
        else:
            try:
                returnValue = element.text
            except:
                returnValue = element.get_attribute("innerText")

        return returnValue

    def get_element(self, selector_type, selector):
        # print(f"in get_element selector_type = {selector_type}")
        if selector_type.lower() == "xpath":
            return self.driver.find_element(By.XPATH, selector)
        elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
            return self.driver.find_element(By.CSS_SELECTOR, selector)
        elif selector_type.lower() == "id":
            return self.driver.find_element(By.ID, selector)
        elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
            return self.driver.find_element(By.TAG_NAME, selector)
        elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
            return self.driver.find_element(By.LINK_TEXT, selector)
        elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
            return self.driver.find_element(By.CLASS_NAME, selector)
        else:
            print(f"in get_element else statement: selector_type = {selector_type}")
            return self.driver.find_element(By.NAME, selector)

    def get_elements(self, selector_type, selector):
        # print(f"in get_element selector_type = {selector_type}")
        if selector_type.lower() == "xpath":
            return self.driver.find_elements(By.XPATH, selector)
        elif selector_type.lower() == "cssselector" or selector_type.lower() == "css_selector":
            return self.driver.find_elements(By.CSS_SELECTOR, selector)
        elif selector_type.lower() == "id":
            return self.driver.find_elements(By.ID, selector)
        elif selector_type.lower() == "tagname" or selector_type.lower() == "tag_name":
            return self.driver.find_elements(By.TAG_NAME, selector)
        elif selector_type.lower() == "linktext" or selector_type.lower() == "link_text":
            return self.driver.find_elements(By.LINK_TEXT, selector)
        elif selector_type.lower() == "classname" or selector_type.lower() == "class_name":
            return self.driver.find_elements(By.CLASS_NAME, selector)
        else:
            print(f"in get_element else statement: selector_type = {selector_type}")
            return self.driver.find_elements(By.NAME, selector)

    def log_equal_action(self, action, expected, actual, description):
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        # print_line(f"In log_equal_action workbook = {workbook.path}, sheet = {sheet.title}")
        status = "Fail"
        # print(f"In log_equal_action action = {action},expected = {expected}, actual = {actual}, description = {description}")
        now = datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        if expected == actual:
            status = "Pass"
        row = self.find_first_empty_row()
        if row == 1:
            self.print_headers(sheet)
            row = row + 1

        sheet.cell(row, 1).value = action
        sheet.cell(row, 2).value = expected
        sheet.cell(row, 3).value = actual
        sheet.cell(row, 4).value = status
        sheet.cell(row, 5).value = str(now)
        sheet.cell(row, 6).value = description

        if status == "Pass":
            sheet.cell(row, 4).style = 'Good'
        else:
            sheet.cell(row, 4).style = 'Bad'

        # need to account for self.driver being unavailable because quit() was already initiated
        # if self.driver.session_id is None:
        if action.lower() == "close driver":
            sheet.cell(row, 7).value = "n/a"
            self.set_close_driver_theme(sheet, row)
        else:
            sheet.cell(row, 7).value = self.driver.current_url
        workbook.save(self.log_file_name)
        print(f"Action: {action}\t|\tDescription:{description}\t|\tStatus:{status}")

    def log_contains_action(self, action, expected, actual, description):
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        now = datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S')
        status = self.get_status_contains(expected, actual)
        row = self.find_first_empty_row()
        if row == 1:
            self.print_headers(sheet)
            row = row + 1
        sheet.cell(row, 1).value = action
        sheet.cell(row, 2).value = expected
        sheet.cell(row, 3).value = actual
        sheet.cell(row, 4).value = status
        sheet.cell(row, 5).value = str(now)
        sheet.cell(row, 6).value = description

        if status == "Pass":
            sheet.cell(row, 4).style = 'Good'

        else:
            sheet.cell(row, 4).style = 'Bad'
        # need to account for self.driver being unavailable because quit() was already initiated
        # if self.driver.session_id is None:
        if action.lower() == "close driver":
            print("...")
            sheet.cell(row, 7).value = "n/a"
            self.set_close_driver_theme(sheet, row)
            print(" \n ")
        else:
            sheet.cell(row, 7).value = self.driver.current_url
        workbook.save(self.log_file_name)
        print(f"Action: {action}\t|\tDescription:{description}\t|\tStatus:{status}")

    # This method finds the first empty row in the Excel Spreadsheet
    def find_first_empty_row(self):
        workbook = openpyxl.load_workbook(self.log_file_name)
        sheet = workbook[self.log_sheet_name]
        row = 1
        while sheet.cell(row, 1).value is not None and len(sheet.cell(row, 1).value) > 0:
            row = row + 1

        return row


    # This method prints the header titles if headers have not been written
    @staticmethod
    def print_headers(sheet):
        row = 1
        start_style = "Neutral"
        sheet.cell(row, 1).value = "Action"
        sheet.cell(row, 2).value = "Expected"
        sheet.cell(row, 3).value = "Actual"
        sheet.cell(row, 4).value = "Status"
        sheet.cell(row, 5).value = "Date and Time Executed"
        sheet.cell(row, 6).value = "Description"
        sheet.cell(row, 7).value = "Page URL"
        sheet.cell(row, 1).style = start_style
        sheet.cell(row, 2).style = start_style
        sheet.cell(row, 3).style = start_style
        sheet.cell(row, 4).style = start_style
        sheet.cell(row, 5).style = start_style
        sheet.cell(row, 6).style = start_style
        sheet.cell(row, 7).style = start_style

    def set_close_driver_theme(self, sheet, row):
        end_style = "Neutral"
        sheet.cell(row, 1).style = end_style
        sheet.cell(row, 2).style = end_style
        sheet.cell(row, 3).style = end_style
        sheet.cell(row, 4).style = end_style
        sheet.cell(row, 5).style = end_style
        sheet.cell(row, 6).style = end_style
        sheet.cell(row, 7).style = end_style

    @staticmethod
    def get_status_contains(contained_string, full_string):
        print(f"contained_string = {contained_string}")
        print(f"full_string = {full_string}")
        if contained_string in full_string:
            return "Pass"
        else:
            return "Fail"

    def get_json_from_api(self, api_url):
        # requests.get(self.driver.current_url).json()
        print("in get_json_from_api")
        json_data = requests.get(api_url).json()
        min_name_len = 35
        min_genre_len = 17
        spaces = (min_name_len - len('name')) * "."
        g_spaces = (min_genre_len - len('genre')) * "."
        print(f"Id\tName{spaces}\tGenre{g_spaces}\tPrice\tRelease Date")
        for game in json_data:
            spaces = (min_name_len - len(game['name']))* "."
            g_spaces = (min_genre_len - len(game['genre']))* "."
            print(f"{game['id']}\t{game['name']}{spaces}\t{game['genre']}{g_spaces}\t,{game['price']}\t{game['releaseDate']}\n")